# app.py - Complete Working Version with Full Authentication System & Excel Templates
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, jsonify
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import pandas as pd
import numpy as np
import os
import sqlite3
from datetime import datetime
import io
import csv
import json
from functools import wraps
import re
import zipfile

app = Flask(__name__)
app.secret_key = 'sales-analytics-dashboard-secret-key-2024-prod'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['ALLOWED_EXTENSIONS'] = {'csv', 'xls', 'xlsx'}
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Database setup
DB_NAME = 'sales.db'

def get_db():
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    cursor = conn.cursor()
    
    # Drop existing tables if they exist
    cursor.execute('DROP TABLE IF EXISTS sales')
    cursor.execute('DROP TABLE IF EXISTS users')
    
    # Create sales table
    cursor.execute('''
    CREATE TABLE sales (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_id TEXT,
        order_date TEXT,
        customer_name TEXT,
        product_name TEXT,
        category TEXT,
        quantity REAL,
        unit_price REAL,
        total_price REAL,
        country TEXT,
        payment_mode TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    ''')
    
    # Create indexes for better performance
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_order_date ON sales(order_date)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_category ON sales(category)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_customer ON sales(customer_name)')
    
    # Create users table with additional fields
    cursor.execute('''
    CREATE TABLE users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        email TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL,
        full_name TEXT,
        role TEXT DEFAULT 'user',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        last_login TIMESTAMP,
        is_active INTEGER DEFAULT 1
    )
    ''')
    
    # Insert default admin user with HASHED password
    hashed_password = generate_password_hash('1234')
    cursor.execute("""
        INSERT OR IGNORE INTO users (email, password, full_name, role) 
        VALUES (?, ?, ?, ?)
    """, ('admin@example.com', hashed_password, 'Administrator', 'admin'))
    
    # Insert default test user
    test_password = generate_password_hash('test123')
    cursor.execute("""
        INSERT OR IGNORE INTO users (email, password, full_name) 
        VALUES (?, ?, ?)
    """, ('user@example.com', test_password, 'Test User'))
    
    conn.commit()
    conn.close()
    print("✅ Database initialized successfully with default users")

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session:
            flash('Please login first', 'danger')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session:
            flash('Please login first', 'danger')
            return redirect(url_for('login'))
        
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT role FROM users WHERE email = ?", (session['user'],))
        user = cursor.fetchone()
        conn.close()
        
        if not user or user['role'] != 'admin':
            flash('Admin access required', 'danger')
            return redirect(url_for('dashboard'))
        
        return f(*args, **kwargs)
    return decorated_function

# Template Design Functions
def apply_simple_design(worksheet, df):
    """Simple and clean design"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    # Simple header
    header_font = Font(name='Arial', size=11, bold=True, color='000000')
    header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    
    # Apply to headers
    for col_num in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in worksheet.iter_rows(min_row=1, max_row=len(df)+1, max_col=len(df.columns)):
        for cell in row:
            cell.border = thin_border
    
    # Auto-adjust column widths
    from openpyxl.utils import get_column_letter
    
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 30)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def apply_modern_design(worksheet, df):
    """Modern gradient design"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
    
    # Modern gradient header
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='2E75B6', fill_type='solid')
    
    # Apply to headers
    for col_num in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Zebra striping for data rows
    for row_num in range(2, len(df) + 2):
        fill_color = 'FFFFFF' if row_num % 2 == 0 else 'F2F2F2'
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.fill = row_fill
    
    # Modern border
    modern_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    for row in worksheet.iter_rows(min_row=1, max_row=len(df)+1, max_col=len(df.columns)):
        for cell in row:
            cell.border = modern_border
    
    # Auto-adjust column widths
    from openpyxl.utils import get_column_letter
    
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 35)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def apply_colorful_design(worksheet, df):
    """Colorful category-based design"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    # Color palette for different categories
    color_palette = {
        'Order ID': 'FF6B6B',      # Red
        'Order Date': '4ECDC4',    # Teal
        'Customer Name': '45B7D1', # Blue
        'Product Name': '96CEB4',  # Green
        'Category': 'FFEAA7',      # Yellow
        'Quantity': 'DDA0DD',      # Purple
        'Unit Price': 'FFA07A',    # Orange
        'Total Price': '98D8C8',   # Mint
        'Country': 'F7DC6F',       # Gold
        'Payment Mode': 'BB8FCE'   # Lavender
    }
    
    # Apply colorful headers
    for col_num, column_title in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=color_palette.get(column_title, '4F81BD'), 
                              end_color=color_palette.get(column_title, '4F81BD'), 
                              fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Colorful alternating rows
    for row_num in range(2, len(df) + 2):
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            
            # Light version of header color for data cells
            header_color = color_palette.get(column_title, '4F81BD')
            light_color = header_color
            
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            else:
                cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
    
    # Rounded corner style borders
    rounded_border = Border(
        left=Side(style='medium', color='666666'),
        right=Side(style='medium', color='666666'),
        top=Side(style='medium', color='666666'),
        bottom=Side(style='medium', color='666666')
    )
    
    for row in worksheet.iter_rows(min_row=1, max_row=len(df)+1, max_col=len(df.columns)):
        for cell in row:
            cell.border = rounded_border
    
    # Wider columns for better visibility
    from openpyxl.utils import get_column_letter
    
    for column in worksheet.columns:
        column_letter = get_column_letter(column[0].column)
        worksheet.column_dimensions[column_letter].width = 20

def add_instructions_sheet(writer, df):
    """Add instructions sheet to workbook"""
    instructions_data = {
        'Column': list(df.columns),
        'Description': [
            'Unique order identifier (required)',
            'Date in YYYY-MM-DD format (required)',
            'Customer name (required)',
            'Product name (required)',
            'Product category (optional)',
            'Quantity sold (required, numeric)',
            'Price per unit (required, numeric)',
            'Total amount (required, numeric)',
            'Customer country (optional)',
            'Payment method (optional)'
        ],
        'Format': [
            'Text (ORD_100001, 1001, TRANS-001)',
            'Date (2024-01-15, 15/01/2024)',
            'Text (John Smith, Customer A)',
            'Text (Laptop, Product XYZ)',
            'Text (Electronics, Clothing)',
            'Number (1, 2.5, 10)',
            'Number (75000.00, 99.99)',
            'Number (75000.00, 199.98)',
            'Text (India, USA, UK)',
            'Text (Credit Card, PayPal)'
        ],
        'Required': ['Yes', 'Yes', 'Yes', 'Yes', 'No', 'Yes', 'Yes', 'Yes', 'No', 'No']
    }
    
    instructions_df = pd.DataFrame(instructions_data)
    instructions_df.to_excel(writer, sheet_name='Instructions', index=False)
    
    # Format instructions sheet
    ws_instructions = writer.sheets['Instructions']
    
    # Format headers
    from openpyxl.styles import Font, PatternFill, Alignment
    
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    
    for col_num in range(1, len(instructions_df.columns) + 1):
        cell = ws_instructions.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Auto-adjust column widths
    from openpyxl.utils import get_column_letter
    
    for column in ws_instructions.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 40)
        ws_instructions.column_dimensions[column_letter].width = adjusted_width

def add_validation_sheet(writer):
    """Add validation rules and examples sheet"""
    validation_data = {
        'Rule Type': ['Data Format', 'Required Fields', 'Numeric Values', 'Date Format', 'Text Length'],
        'Description': [
            'Follow the column formats specified',
            'Marked "Yes" in Required column must be filled',
            'Quantity, Unit Price, Total Price must be numbers',
            'Order Date should be in proper date format',
            'Text fields should not exceed 100 characters'
        ],
        'Example': [
            'See Format column in Instructions',
            'Order ID, Order Date, Customer Name, etc.',
            '1, 99.99, 1500.00 (no commas, no currency symbols)',
            '2024-01-15 (YYYY-MM-DD) recommended',
            'Customer Name: John Smith (max 100 chars)'
        ],
        'Error Message': [
            'Invalid format for column',
            'Required field is empty',
            'Numeric value expected',
            'Invalid date format',
            'Text too long'
        ]
    }
    
    validation_df = pd.DataFrame(validation_data)
    validation_df.to_excel(writer, sheet_name='Validation Rules', index=False)
    
    # Format validation sheet
    ws_validation = writer.sheets['Validation Rules']
    
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Format headers
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='D35400', end_color='D35400', fill_type='solid')
    
    for col_num in range(1, len(validation_df.columns) + 1):
        cell = ws_validation.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Auto-adjust column widths
    from openpyxl.utils import get_column_letter
    
    for column in ws_validation.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 35)
        ws_validation.column_dimensions[column_letter].width = adjusted_width

# Template filters
@app.template_filter('format_currency')
def format_currency(value):
    try:
        value = float(value)
        return f"₹{value:,.2f}"
    except:
        return "₹0.00"

@app.template_filter('format_number')
def format_number(value):
    try:
        value = int(value)
        return f"{value:,}"
    except:
        return "0"

@app.template_filter('format_date')
def format_date(value):
    try:
        return datetime.strptime(value, '%Y-%m-%d').strftime('%d %b %Y')
    except:
        return value

@app.template_filter('format_datetime')
def format_datetime(value):
    try:
        return datetime.strptime(value, '%Y-%m-%d %H:%M:%S').strftime('%d %b %Y, %I:%M %p')
    except:
        return value

# Authentication Routes
@app.route('/')
def home():
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    # If user is already logged in, redirect to dashboard
    if 'user' in session:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '').strip()
        remember = request.form.get('remember', False)
        
        # Validate inputs
        errors = []
        
        if not email:
            errors.append('Email is required')
        
        if not password:
            errors.append('Password is required')
        
        # Validate email format
        if email and not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email):
            errors.append('Invalid email format')
        
        if errors:
            for error in errors:
                flash(error, 'danger')
            return render_template('login.html')
        
        conn = get_db()
        cursor = conn.cursor()
        
        try:
            # Get user by email
            cursor.execute("""
                SELECT * FROM users 
                WHERE email = ? AND is_active = 1
            """, (email,))
            user = cursor.fetchone()
            
            # Check if user exists and password matches
            if user and check_password_hash(user['password'], password):
                # Update last login time
                cursor.execute("""
                    UPDATE users 
                    SET last_login = CURRENT_TIMESTAMP 
                    WHERE id = ?
                """, (user['id'],))
                conn.commit()
                
                # Set session variables
                session['user'] = user['email']
                session['username'] = user['email'].split('@')[0]
                session['user_id'] = user['id']
                session['full_name'] = user['full_name'] or user['email'].split('@')[0]
                session['role'] = user['role']
                
                # Set session as permanent if remember me is checked
                if remember:
                    session.permanent = True
                    app.permanent_session_lifetime = 86400  # 24 hours
                else:
                    session.permanent = False
                
                conn.close()
                
                flash('Login successful! Welcome to Sales Analytics Dashboard', 'success')
                print(f"✅ User logged in: {email}")
                return redirect(url_for('dashboard'))
            else:
                conn.close()
                flash('Invalid email or password', 'danger')
                print(f"❌ Failed login attempt: {email}")
                
        except Exception as e:
            conn.close()
            print(f"❌ Database error during login: {e}")
            flash('An error occurred. Please try again.', 'danger')
    
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    # If user is already logged in, redirect to dashboard
    if 'user' in session:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()
        full_name = request.form.get('full_name', '').strip()
        
        # Validate inputs
        errors = []
        
        if not email:
            errors.append('Email is required')
        
        if not password:
            errors.append('Password is required')
        
        if not confirm_password:
            errors.append('Please confirm your password')
        
        if not full_name:
            errors.append('Full name is required')
        
        # Validate email format
        if email and not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email):
            errors.append('Invalid email format')
        
        # Validate password strength
        if password:
            if len(password) < 6:
                errors.append('Password must be at least 6 characters')
            if password != confirm_password:
                errors.append('Passwords do not match')
        
        if errors:
            for error in errors:
                flash(error, 'danger')
            return render_template('register.html', 
                                 email=email, 
                                 full_name=full_name)
        
        conn = get_db()
        cursor = conn.cursor()
        
        try:
            # Check if user already exists
            cursor.execute("SELECT * FROM users WHERE email = ?", (email,))
            existing_user = cursor.fetchone()
            
            if existing_user:
                flash('Email already registered. Please use a different email or login.', 'danger')
                conn.close()
                return render_template('register.html', 
                                     email=email, 
                                     full_name=full_name)
            
            # Create new user with hashed password
            hashed_password = generate_password_hash(password)
            cursor.execute("""
                INSERT INTO users (email, password, full_name, role) 
                VALUES (?, ?, ?, ?)
            """, (email, hashed_password, full_name, 'user'))
            
            conn.commit()
            
            # Get the new user's ID
            user_id = cursor.lastrowid
            
            conn.close()
            
            flash('Registration successful! Please login with your credentials.', 'success')
            print(f"✅ New user registered: {email}")
            return redirect(url_for('login'))
            
        except Exception as e:
            conn.close()
            print(f"❌ Database error during registration: {e}")
            flash('An error occurred during registration. Please try again.', 'danger')
            return render_template('register.html', 
                                 email=email, 
                                 full_name=full_name)
    
    return render_template('register.html')

@app.route('/logout')
def logout():
    if 'user' in session:
        print(f"✅ User logged out: {session['user']}")
    
    session.clear()
    flash('Logged out successfully', 'info')
    return redirect(url_for('login'))

@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    conn = get_db()
    cursor = conn.cursor()
    
    if request.method == 'POST':
        full_name = request.form.get('full_name', '').strip()
        
        if not full_name:
            flash('Full name is required', 'danger')
        else:
            try:
                cursor.execute("""
                    UPDATE users 
                    SET full_name = ? 
                    WHERE id = ?
                """, (full_name, session['user_id']))
                conn.commit()
                
                session['full_name'] = full_name
                flash('Profile updated successfully!', 'success')
                print(f"✅ Profile updated for user: {session['user']}")
                
            except Exception as e:
                print(f"❌ Error updating profile: {e}")
                flash('An error occurred. Please try again.', 'danger')
    
    # Get current user data
    cursor.execute("""
        SELECT email, full_name, role, created_at, last_login 
        FROM users 
        WHERE id = ?
    """, (session['user_id'],))
    user = cursor.fetchone()
    
    conn.close()
    
    return render_template('profile.html', 
                         user=user,
                         username=session.get('username', 'User'),
                         full_name=session.get('full_name', ''))

@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        current_password = request.form.get('current_password', '').strip()
        new_password = request.form.get('new_password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()
        
        # Validate inputs
        errors = []
        
        if not current_password:
            errors.append('Current password is required')
        
        if not new_password:
            errors.append('New password is required')
        
        if not confirm_password:
            errors.append('Please confirm your new password')
        
        if new_password and len(new_password) < 6:
            errors.append('New password must be at least 6 characters')
        
        if new_password and confirm_password and new_password != confirm_password:
            errors.append('New passwords do not match')
        
        if errors:
            for error in errors:
                flash(error, 'danger')
            return render_template('change_password.html', 
                                 username=session.get('username', 'User'))
        
        conn = get_db()
        cursor = conn.cursor()
        
        try:
            # Get current user with password
            cursor.execute("SELECT password FROM users WHERE id = ?", (session['user_id'],))
            user = cursor.fetchone()
            
            # Verify current password
            if user and check_password_hash(user['password'], current_password):
                # Update to new password
                hashed_password = generate_password_hash(new_password)
                cursor.execute("""
                    UPDATE users 
                    SET password = ? 
                    WHERE id = ?
                """, (hashed_password, session['user_id']))
                conn.commit()
                conn.close()
                
                flash('Password changed successfully!', 'success')
                print(f"✅ Password changed for user: {session['user']}")
                return redirect(url_for('dashboard'))
            else:
                conn.close()
                flash('Current password is incorrect', 'danger')
                
        except Exception as e:
            conn.close()
            print(f"❌ Error changing password: {e}")
            flash('An error occurred. Please try again.', 'danger')
    
    return render_template('change_password.html', 
                         username=session.get('username', 'User'))

# Template Download Routes
@app.route('/download-template/<template_type>')
@login_required
def download_template(template_type):
    """Download Excel template for data upload with different designs"""
    try:
        # Sample data common for all templates
        sample_data = {
            'Order ID': ['ORD_100001', 'ORD_100002', 'ORD_100003', 'ORD_100004', 'ORD_100005'],
            'Order Date': ['2024-01-15', '2024-01-16', '2024-01-17', '2024-01-18', '2024-01-19'],
            'Customer Name': ['John Smith', 'Jane Doe', 'Robert Johnson', 'Emily Williams', 'Michael Brown'],
            'Product Name': ['Laptop Pro', 'Wireless Mouse', 'Keyboard', 'Monitor 24"', 'Webcam HD'],
            'Category': ['Electronics', 'Accessories', 'Accessories', 'Electronics', 'Accessories'],
            'Quantity': [1, 2, 1, 1, 3],
            'Unit Price': [75000.00, 1200.00, 2500.00, 15000.00, 1800.00],
            'Total Price': [75000.00, 2400.00, 2500.00, 15000.00, 5400.00],
            'Country': ['India', 'USA', 'UK', 'Canada', 'Australia'],
            'Payment Mode': ['Credit Card', 'PayPal', 'Debit Card', 'Credit Card', 'UPI']
        }
        
        df = pd.DataFrame(sample_data)
        
        # Create Excel file in memory
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Sales Data', index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Sales Data']
            
            # Apply different designs based on template type
            if template_type == 'simple':
                apply_simple_design(worksheet, df)
            elif template_type == 'modern':
                apply_modern_design(worksheet, df)
            elif template_type == 'colorful':
                apply_colorful_design(worksheet, df)
            else:
                # Default design
                apply_simple_design(worksheet, df)
            
            # Add instructions sheet for all templates
            add_instructions_sheet(writer, df)
            
            # Add data validation rules sheet
            add_validation_sheet(writer)
        
        output.seek(0)
        
        # Set filename
        filename = f"sales_template_{template_type}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"❌ Template generation error: {e}")
        flash(f'Error generating template: {str(e)}', 'danger')
        return redirect(url_for('upload'))

@app.route('/templates')
@login_required
def templates():
    """Template download page"""
    return render_template('templates.html', 
                         username=session.get('username', 'User'),
                         full_name=session.get('full_name', ''))

@app.route('/download-all-templates')
@login_required
def download_all_templates():
    """Download all templates as ZIP file"""
    try:
        # Create ZIP file in memory
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Generate and add each template
            template_types = ['simple', 'modern', 'colorful']
            
            for template_type in template_types:
                # Generate each template
                sample_data = {
                    'Order ID': ['ORD_100001', 'ORD_100002'],
                    'Order Date': ['2024-01-15', '2024-01-16'],
                    'Customer Name': ['John Smith', 'Jane Doe'],
                    'Product Name': ['Laptop Pro', 'Wireless Mouse'],
                    'Category': ['Electronics', 'Accessories'],
                    'Quantity': [1, 2],
                    'Unit Price': [75000.00, 1200.00],
                    'Total Price': [75000.00, 2400.00],
                    'Country': ['India', 'USA'],
                    'Payment Mode': ['Credit Card', 'PayPal']
                }
                
                df = pd.DataFrame(sample_data)
                excel_buffer = io.BytesIO()
                
                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Sales Data', index=False)
                    
                    worksheet = writer.sheets['Sales Data']
                    if template_type == 'simple':
                        apply_simple_design(worksheet, df)
                    elif template_type == 'modern':
                        apply_modern_design(worksheet, df)
                    elif template_type == 'colorful':
                        apply_colorful_design(worksheet, df)
                    
                    add_instructions_sheet(writer, df)
                    add_validation_sheet(writer)
                
                excel_buffer.seek(0)
                zip_file.writestr(f'sales_template_{template_type}.xlsx', excel_buffer.getvalue())
        
        zip_buffer.seek(0)
        
        return send_file(
            zip_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name=f'sales_templates_{datetime.now().strftime("%Y%m%d")}.zip'
        )
        
    except Exception as e:
        print(f"❌ ZIP generation error: {e}")
        flash('Error generating template package', 'danger')
        return redirect(url_for('templates'))

# Admin routes for user management
@app.route('/admin/users')
@admin_required
def admin_users():
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT id, email, full_name, role, created_at, last_login, is_active
        FROM users 
        ORDER BY created_at DESC
    """)
    users = cursor.fetchall()
    
    conn.close()
    
    return render_template('admin_users.html',
                         users=users,
                         username=session.get('username', 'User'),
                         role=session.get('role', 'user'))

# Main Application Routes
@app.route('/dashboard')
@login_required
def dashboard():
    conn = get_db()
    cursor = conn.cursor()
    
    # Check if we have data
    cursor.execute("SELECT COUNT(*) FROM sales")
    count = cursor.fetchone()[0]
    
    if count == 0:
        conn.close()
        return render_template('dashboard.html', 
                             has_data=False,
                             username=session.get('username', 'User'),
                             full_name=session.get('full_name', ''),
                             now=datetime.now())
    
    # Get KPIs
    cursor.execute('''
    SELECT 
        COUNT(DISTINCT order_id) as total_orders,
        COUNT(DISTINCT customer_name) as total_customers,
        SUM(total_price) as total_revenue,
        AVG(total_price) as avg_order_value,
        SUM(quantity) as total_quantity
    FROM sales
    ''')
    kpi_row = cursor.fetchone()
    
    kpis = {
        'total_orders': kpi_row[0] or 0,
        'total_customers': kpi_row[1] or 0,
        'total_revenue': kpi_row[2] or 0,
        'avg_order_value': kpi_row[3] or 0,
        'total_quantity': kpi_row[4] or 0
    }
    
    # Get top categories
    cursor.execute('''
    SELECT category, SUM(total_price) as revenue
    FROM sales
    WHERE category IS NOT NULL AND category != ''
    GROUP BY category
    ORDER BY revenue DESC
    LIMIT 5
    ''')
    top_categories = [{'category': row[0], 'revenue': row[1]} for row in cursor.fetchall()]
    
    # Get recent sales
    cursor.execute('''
    SELECT order_id, order_date, customer_name, product_name, total_price
    FROM sales
    ORDER BY order_date DESC, created_at DESC
    LIMIT 10
    ''')
    recent_sales = cursor.fetchall()
    
    conn.close()
    
    return render_template('dashboard.html',
                         has_data=True,
                         kpis=kpis,
                         top_categories=top_categories,
                         recent_sales=recent_sales,
                         now=datetime.now(),
                         username=session.get('username', 'User'),
                         full_name=session.get('full_name', ''))

@app.route('/api/chart-data')
@login_required
def api_chart_data():
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Revenue by category
        cursor.execute('''
        SELECT category, SUM(total_price) as revenue
        FROM sales
        WHERE category != 'Unknown'
        GROUP BY category
        ORDER BY revenue DESC
        LIMIT 10
        ''')
        revenue_by_category = {}
        for row in cursor.fetchall():
            revenue_by_category[row[0]] = float(row[1] or 0)
        
        # Monthly sales trend (last 6 months)
        cursor.execute('''
        SELECT 
            strftime('%Y-%m', order_date) as month,
            SUM(total_price) as revenue
        FROM sales
        WHERE order_date IS NOT NULL AND order_date != ''
        GROUP BY strftime('%Y-%m', order_date)
        ORDER BY month DESC
        LIMIT 6
        ''')
        monthly_sales = {}
        for row in cursor.fetchall():
            if row[0]:
                try:
                    month_name = datetime.strptime(row[0] + '-01', '%Y-%m-%d').strftime('%b %Y')
                    monthly_sales[month_name] = float(row[1] or 0)
                except:
                    pass
        
        # Country revenue
        cursor.execute('''
        SELECT country, SUM(total_price) as revenue
        FROM sales
        WHERE country != 'Unknown'
        GROUP BY country
        ORDER BY revenue DESC
        LIMIT 10
        ''')
        country_revenue = {}
        for row in cursor.fetchall():
            country_revenue[row[0]] = float(row[1] or 0)
        
        # Payment distribution
        cursor.execute('''
        SELECT payment_mode, COUNT(*) as count
        FROM sales
        WHERE payment_mode != 'Unknown'
        GROUP BY payment_mode
        ''')
        payment_distribution = {}
        for row in cursor.fetchall():
            payment_distribution[row[0]] = int(row[1] or 0)
        
        # Daily sales for last 30 days
        cursor.execute('''
        SELECT 
            date(order_date) as day,
            SUM(total_price) as revenue
        FROM sales
        WHERE order_date >= date('now', '-30 days')
        GROUP BY date(order_date)
        ORDER BY day
        ''')
        daily_sales = {}
        for row in cursor.fetchall():
            if row[0]:
                try:
                    day_name = datetime.strptime(row[0], '%Y-%m-%d').strftime('%d %b')
                    daily_sales[day_name] = float(row[1] or 0)
                except:
                    pass
        
        conn.close()
        
        return jsonify({
            'revenue_by_category': revenue_by_category,
            'monthly_sales': monthly_sales,
            'country_revenue': country_revenue,
            'payment_distribution': payment_distribution,
            'daily_sales': daily_sales
        })
        
    except Exception as e:
        conn.close()
        print(f"❌ Chart data error: {e}")
        return jsonify({'error': 'Failed to fetch chart data'}), 500

@app.route('/upload', methods=['GET', 'POST'])
@login_required
def upload():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file selected', 'danger')
            return redirect(request.url)
        
        file = request.files['file']
        
        if file.filename == '':
            flash('No file selected', 'danger')
            return redirect(request.url)
        
        # Validate file type
        if not allowed_file(file.filename):
            flash('Invalid file type. Please upload CSV or Excel files only.', 'danger')
            return redirect(request.url)
        
        try:
            # Secure filename
            filename = secure_filename(file.filename)
            file_ext = filename.rsplit('.', 1)[1].lower()
            
            # Read file
            if file_ext == 'csv':
                df = pd.read_csv(file)
            else:
                df = pd.read_excel(file)
            
            print(f"📊 File loaded: {df.shape[0]} rows, {df.shape[1]} columns")
            
            # Clean column names
            df.columns = [str(col).strip().lower().replace(' ', '_').replace('-', '_') 
                         for col in df.columns]
            
            # Common column mappings
            column_mappings = {
                'order_id': ['order_id', 'orderid', 'id', 'transaction_id'],
                'order_date': ['order_date', 'date', 'purchase_date', 'transaction_date'],
                'customer_name': ['customer_name', 'customer', 'client_name', 'buyer'],
                'product_name': ['product_name', 'product', 'item', 'item_name'],
                'category': ['category', 'product_category', 'type', 'product_type'],
                'quantity': ['quantity', 'qty', 'units', 'number'],
                'unit_price': ['unit_price', 'price', 'unit_cost', 'cost'],
                'total_price': ['total_price', 'total', 'amount', 'revenue', 'sales'],
                'country': ['country', 'customer_country', 'location'],
                'payment_mode': ['payment_mode', 'payment_method', 'payment', 'method']
            }
            
            # Map columns
            for standard_name, possible_names in column_mappings.items():
                for name in possible_names:
                    if name in df.columns:
                        df[standard_name] = df[name]
                        break
            
            # Ensure required columns exist
            required_cols = ['order_id', 'order_date', 'customer_name', 'product_name', 
                           'quantity', 'unit_price', 'total_price']
            
            for col in required_cols:
                if col not in df.columns:
                    if col == 'order_id':
                        df[col] = ['ORD_' + str(i+1).zfill(6) for i in range(len(df))]
                    else:
                        df[col] = None
            
            # Clean data
            # Dates
            if 'order_date' in df.columns:
                df['order_date'] = pd.to_datetime(df['order_date'], errors='coerce')
                df['order_date'] = df['order_date'].dt.strftime('%Y-%m-%d')
            
            # Numeric columns
            numeric_cols = ['quantity', 'unit_price', 'total_price']
            for col in numeric_cols:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
            
            # Text columns
            text_cols = ['customer_name', 'product_name', 'category', 'country', 'payment_mode']
            for col in text_cols:
                if col in df.columns:
                    df[col] = df[col].astype(str).str.strip().fillna('Unknown')
                    df[col] = df[col].replace(['nan', 'NaN', 'None', '', 'null', 'NULL'], 'Unknown')
            
            # Validate data
            if len(df) == 0:
                flash('Uploaded file contains no valid data', 'danger')
                return redirect(request.url)
            
            # Check for duplicate orders
            conn = get_db()
            cursor = conn.cursor()
            
            # Get existing order IDs
            cursor.execute("SELECT DISTINCT order_id FROM sales")
            existing_orders = set(row[0] for row in cursor.fetchall())
            
            # Filter out duplicates
            new_data = []
            duplicate_count = 0
            
            for _, row in df.iterrows():
                order_id = str(row.get('order_id', ''))
                if order_id not in existing_orders:
                    new_data.append((
                        order_id,
                        str(row.get('order_date', '')),
                        str(row.get('customer_name', 'Unknown'))[:100],
                        str(row.get('product_name', 'Unknown'))[:100],
                        str(row.get('category', 'Unknown'))[:50],
                        float(row.get('quantity', 0)),
                        float(row.get('unit_price', 0)),
                        float(row.get('total_price', 0)),
                        str(row.get('country', 'Unknown'))[:50],
                        str(row.get('payment_mode', 'Unknown'))[:50]
                    ))
                else:
                    duplicate_count += 1
            
            # Batch insert
            if new_data:
                cursor.executemany('''
                    INSERT INTO sales (order_id, order_date, customer_name, product_name, 
                                      category, quantity, unit_price, total_price, 
                                      country, payment_mode)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', new_data)
                
                conn.commit()
                inserted_count = len(new_data)
                
                if duplicate_count > 0:
                    flash(f'Successfully uploaded {inserted_count} records. {duplicate_count} duplicate records were skipped.', 'warning')
                else:
                    flash(f'Successfully uploaded {inserted_count} records!', 'success')
                
                print(f"✅ Uploaded {inserted_count} records from {filename}")
            else:
                flash('No new records to upload. All records already exist in database.', 'warning')
            
            conn.close()
            
            return redirect(url_for('dashboard'))
            
        except pd.errors.EmptyDataError:
            flash('The uploaded file is empty', 'danger')
            return redirect(request.url)
        except pd.errors.ParserError:
            flash('Invalid file format. Please check the file.', 'danger')
            return redirect(request.url)
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"❌ Upload error: {str(e)}\n{error_details}")
            flash(f'Error processing file: {str(e)}', 'danger')
            return redirect(request.url)
    
    return render_template('upload.html', 
                         username=session.get('username', 'User'),
                         full_name=session.get('full_name', ''))

@app.route('/export/csv')
@login_required
def export_csv():
    conn = get_db()
    df = pd.read_sql_query('SELECT * FROM sales', conn)
    conn.close()
    
    # Format the dataframe
    if 'created_at' in df.columns:
        df = df.drop('created_at', axis=1)
    
    csv_data = df.to_csv(index=False)
    buffer = io.BytesIO()
    buffer.write(csv_data.encode('utf-8'))
    buffer.seek(0)
    
    filename = f"sales_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    
    return send_file(
        buffer,
        mimetype='text/csv',
        as_attachment=True,
        download_name=filename
    )

@app.route('/export/pdf')
@login_required
def export_pdf():
    try:
        # Check if reportlab is available
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.pdfgen import canvas
        except ImportError:
            flash('PDF export requires reportlab package. Install with: pip install reportlab', 'warning')
            return redirect(url_for('dashboard'))
        
        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter
        
        # Header
        p.setFont("Helvetica-Bold", 16)
        p.drawString(100, height - 100, "Sales Analytics Report")
        
        p.setFont("Helvetica", 12)
        p.drawString(100, height - 130, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        p.drawString(100, height - 150, f"Generated by: {session.get('full_name', session.get('username', 'User'))}")
        
        # Get data
        conn = get_db()
        cursor = conn.cursor()
        
        # KPIs
        cursor.execute('''
        SELECT 
            COUNT(DISTINCT order_id) as total_orders,
            COUNT(DISTINCT customer_name) as total_customers,
            SUM(total_price) as total_revenue,
            AVG(total_price) as avg_order_value,
            SUM(quantity) as total_quantity
        FROM sales
        ''')
        kpi_row = cursor.fetchone()
        
        # KPIs Section
        y_pos = height - 200
        p.setFont("Helvetica-Bold", 14)
        p.drawString(100, y_pos, "Key Performance Indicators")
        y_pos -= 30
        
        p.setFont("Helvetica", 12)
        p.drawString(100, y_pos, f"Total Orders: {kpi_row[0] or 0:,}")
        y_pos -= 20
        p.drawString(100, y_pos, f"Total Revenue: ₹{kpi_row[2] or 0:,.2f}")
        y_pos -= 20
        p.drawString(100, y_pos, f"Total Customers: {kpi_row[1] or 0:,}")
        y_pos -= 20
        p.drawString(100, y_pos, f"Average Order Value: ₹{kpi_row[3] or 0:,.2f}")
        y_pos -= 20
        p.drawString(100, y_pos, f"Total Quantity Sold: {kpi_row[4] or 0:,}")
        y_pos -= 40
        
        # Top categories
        cursor.execute('''
        SELECT category, SUM(total_price) as revenue
        FROM sales
        WHERE category != 'Unknown'
        GROUP BY category
        ORDER BY revenue DESC
        LIMIT 5
        ''')
        
        p.setFont("Helvetica-Bold", 14)
        p.drawString(100, y_pos, "Top Categories by Revenue")
        y_pos -= 30
        
        p.setFont("Helvetica", 12)
        for row in cursor.fetchall():
            p.drawString(120, y_pos, f"• {row[0]}: ₹{row[1]:,.2f}")
            y_pos -= 20
        
        conn.close()
        
        # Save PDF
        p.showPage()
        p.save()
        
        buffer.seek(0)
        filename = f"sales_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"PDF generation error: {e}")
        flash(f'Error generating PDF: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))

@app.route('/clear-data', methods=['POST'])
@login_required
def clear_data():
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Get count before deletion
        cursor.execute("SELECT COUNT(*) FROM sales")
        count_before = cursor.fetchone()[0]
        
        # Delete all sales data
        cursor.execute('DELETE FROM sales')
        
        # Reset autoincrement counter
        cursor.execute('DELETE FROM sqlite_sequence WHERE name="sales"')
        
        conn.commit()
        conn.close()
        
        flash(f'Successfully cleared {count_before} sales records!', 'success')
        print(f"✅ Cleared {count_before} sales records")
        
    except Exception as e:
        print(f"❌ Error clearing data: {e}")
        flash('Error clearing data', 'danger')
    
    return redirect(url_for('dashboard'))

# Error handlers
@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_error(e):
    print(f"❌ 500 Error: {e}")
    return render_template('500.html'), 500

@app.errorhandler(413)
def request_entity_too_large(e):
    flash('File size exceeds the 16MB limit', 'danger')
    return redirect(url_for('upload'))

if __name__ == '__main__':
    init_db()
    print("=" * 60)
    print("🚀 Starting Sales Analytics Dashboard")
    print("=" * 60)
    print("📊 Dashboard URL: http://localhost:5000")
    print("👤 Default Admin: admin@example.com / 1234")
    print("👤 Test User: user@example.com / test123")
    print("📁 Upload folder: ./uploads")
    print("🗄️  Database: ./sales.db")
    print("📊 Templates: http://localhost:5000/templates")
    print("=" * 60)
    
    # Create necessary directories
    os.makedirs('templates', exist_ok=True)
    os.makedirs('static/css', exist_ok=True)
    os.makedirs('static/js', exist_ok=True)
    os.makedirs('static/images', exist_ok=True)
    
    app.run(debug=True, host='0.0.0.0', port=5000)